from __future__ import annotations

import dataclasses
from pathlib import Path
from typing import Iterable

from django.db import transaction
from openpyxl import load_workbook

from voters.models import Voter


@dataclasses.dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    total_rows: int = 0
    errors: list[str] = dataclasses.field(default_factory=list)


def _extract_arabic_row(row: Iterable):
    (
        seq,
        party,
        voter_no,
        name,
        mother_name,
        birth_year,
        province,
        correct_name,
    ) = row
    voter_number = str(voter_no).strip() if voter_no else ""
    full_name = str(correct_name or name or "").strip()
    return {
        "voter_number": voter_number,
        "full_name": full_name or (f"Voter {voter_number}" if voter_number else ""),
        "birth_year": str(birth_year or "").strip(),
        "notes": _build_notes(party, mother_name, province, name, correct_name),
    }


def _build_notes(
    party: str | None,
    mother_name: str | None,
    province: str | None,
    original_name: str | None,
    corrected_name: str | None,
) -> str:
    parts: list[str] = []
    if party:
        parts.append(f"الحزب/التحالف: {party}")
    if mother_name:
        parts.append(f"اسم الأم: {mother_name}")
    if province:
        parts.append(f"المحافظة: {province}")
    if original_name and corrected_name and original_name.strip() != corrected_name.strip():
        parts.append(f"الاسم الأصلي: {original_name}")
    return " | ".join(parts)


def import_voters_from_excel(
    excel_path: Path,
    *,
    sheet_name: str | None = None,
    dry_run: bool = False,
) -> ImportResult:
    workbook = load_workbook(excel_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    result = ImportResult()

    def _apply():
        nonlocal result
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            result.total_rows += 1
            data = _extract_arabic_row(row)

            voter_number = data["voter_number"]
            if not voter_number:
                result.errors.append(
                    f"Row {result.total_rows + 2}: missing voter number, skipped."
                )
                continue

            full_name = data["full_name"]
            birth_year_raw = data["birth_year"]
            birth_year = None
            if birth_year_raw:
                try:
                    birth_year = int(birth_year_raw)
                except (TypeError, ValueError):
                    result.errors.append(
                        f"Row {result.total_rows + 2}: invalid birth year {birth_year_raw!r}."
                    )

            defaults = {
                "full_name": full_name or f"Voter {voter_number}",
                "birth_year": birth_year,
                "notes": data["notes"],
                "is_active": True,
            }

            if dry_run:
                continue

            obj, created = Voter.objects.update_or_create(
                voter_number=voter_number,
                defaults=defaults,
            )
            if created:
                result.created += 1
            else:
                result.updated += 1

    if dry_run:
        _apply()
    else:
        with transaction.atomic():
            _apply()

    return result
